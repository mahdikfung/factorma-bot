import logging
import os
from enum import IntEnum, auto
from pathlib import Path
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, CallbackQueryHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters,
)
from openpyxl import load_workbook
import jdatetime
import psycopg2
import tempfile
import shutil
import subprocess

# ---------- Logging ----------
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ---------- Constants & Paths ----------
TEMPLATE_PATH = Path("/app/فاکتور رسمی مهدی خواجه.xlsx")
OUTPUT_DIR = Path("/app/invoices")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

DB_DSN = os.getenv("DATABASE_URL")
if not DB_DSN:
    raise RuntimeError("DATABASE_URL environment variable not set")

ITEMS_PER_PAGE = 2  # 2 customers per page (per request)

# ---------- Conversation States (IntEnum) ----------
class State(IntEnum):
    MAIN_MENU = auto()
    NEW_INVOICE_MENU = auto()
    ENTER_NEW_CUSTOMER_NAME = auto()
    ENTER_NEW_CUSTOMER_STATE = auto()
    ENTER_NEW_CUSTOMER_CITY = auto()
    ENTER_NEW_CUSTOMER_ADDRESS = auto()
    ENTER_NEW_CUSTOMER_PHONE = auto()
    ENTER_SEARCH_NAME = auto()
    SELECT_EXISTING_CUSTOMER = auto()
    ENTER_DATE_SHAMSI = auto()
    ENTER_ITEM_DESC = auto()
    ENTER_ITEM_QTY = auto()
    ENTER_ITEM_UNIT = auto()
    ENTER_ITEM_PRICE = auto()
    ASK_MORE_ITEMS = auto()
    REVIEW_INVOICE = auto()
    CUSTOMER_LIST = auto()
    CUSTOMER_PAGE = auto()
    VIEW_CUSTOMER = auto()
    EDIT_CUSTOMER_FIELD = auto()
    EDIT_CUSTOMER_VALUE = auto()

# ---------- DB helpers ----------
def get_conn():
    return psycopg2.connect(DB_DSN)

def init_db():
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS customers (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    state TEXT NOT NULL,
                    city TEXT NOT NULL,
                    address TEXT NOT NULL,
                    phone TEXT NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoices (
                    id SERIAL PRIMARY KEY,
                    invoice_number INTEGER UNIQUE NOT NULL,
                    date_shamsi TEXT NOT NULL,
                    customer_id INTEGER NOT NULL REFERENCES customers(id),
                    total_amount INTEGER NOT NULL,
                    file_path TEXT NOT NULL
                )
            """)
            cur.execute("""
                CREATE TABLE IF NOT EXISTS invoice_items (
                    id SERIAL PRIMARY KEY,
                    invoice_id INTEGER NOT NULL REFERENCES invoices(id),
                    description TEXT NOT NULL,
                    quantity REAL NOT NULL,
                    unit TEXT NOT NULL,
                    unit_price INTEGER NOT NULL,
                    total_price INTEGER NOT NULL
                )
            """)
        conn.commit()
    finally:
        conn.close()

def get_next_invoice_number():
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT COALESCE(MAX(id), 0) FROM invoices")
            row = cur.fetchone()
            next_id = (row[0] if row else 0) + 1
            return 10000 + next_id
    finally:
        conn.close()

def create_customer(name, state, city, address, phone):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO customers (name, state, city, address, phone)
                VALUES (%s, %s, %s, %s, %s) RETURNING id
                """,
                (name, state, city, address, phone),
            )
            row = cur.fetchone()
            conn.commit()
            return row[0] if row else None
    finally:
        conn.close()

def search_customers(query):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            if query == "*" or query == "":
                cur.execute("SELECT id, name, state, city, address, phone FROM customers ORDER BY name")
            else:
                cur.execute(
                    """
                    SELECT id, name, state, city, address, phone
                    FROM customers
                    WHERE name ILIKE %s
                    ORDER BY name
                    """,
                    (f"%{query}%",),
                )
            return cur.fetchall()
    finally:
        conn.close()

def get_customer(cust_id):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id, name, state, city, address, phone FROM customers WHERE id=%s",
                (cust_id,),
            )
            row = cur.fetchone()
            if row:
                return {
                    "id": row[0],
                    "name": row[1],
                    "state": row[2],
                    "city": row[3],
                    "address": row[4],
                    "phone": row[5],
                }
            return None
    finally:
        conn.close()

def update_customer(cust_id, name, state, city, address, phone):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                UPDATE customers
                SET name=%s, state=%s, city=%s, address=%s, phone=%s
                WHERE id=%s
                """,
                (name, state, city, address, phone, cust_id),
            )
            conn.commit()
    finally:
        conn.close()

def get_customer_invoices(cust_id):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT invoice_number, date_shamsi, total_amount, file_path
                FROM invoices
                WHERE customer_id=%s
                ORDER BY id DESC
                """,
                (cust_id,),
            )
            return cur.fetchall()
    finally:
        conn.close()

def save_invoice(invoice_number, date_shamsi, customer_id, total, file_path):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO invoices (invoice_number, date_shamsi, customer_id, total_amount, file_path)
                VALUES (%s, %s, %s, %s, %s) RETURNING id
                """,
                (invoice_number, date_shamsi, customer_id, total, file_path),
            )
            row = cur.fetchone()
            conn.commit()
            return row[0] if row else None
    finally:
        conn.close()

def save_items(invoice_id, items):
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            for it in items:
                total = int(it["quantity"] * it["unit_price"])
                cur.execute(
                    """
                    INSERT INTO invoice_items
                    (invoice_id, description, quantity, unit, unit_price, total_price)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (
                        invoice_id,
                        it["description"],
                        it["quantity"],
                        it["unit"],
                        it["unit_price"],
                        total,
                    ),
                )
            conn.commit()
    finally:
        conn.close()

# ---------- PDF generation ----------
def generate_invoice_pdf(data):
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # Buyer info
    ws["B11"] = data["customer_name"]
    ws["D11"] = data.get("customer_state", "")
    ws["F11"] = data.get("customer_city", "")
    ws["H11"] = data.get("customer_address", "")
    ws["F14"] = data["customer_phone"]

    # Date and serial
    ws["F2"] = str(data["invoice_number"])
    ws["B4"] = data["date_shamsi"]

    # Items
    start_row = 16
    for i, item in enumerate(data["items"]):
        r = start_row + i
        ws[f"B{r}"] = i + 1
        ws[f"C{r}"] = item["description"]
        ws[f"D{r}"] = item["quantity"]
        ws[f"E{r}"] = item["unit"]
        ws[f"F{r}"] = item["unit_price"]
        ws[f"G{r}"] = int(item["quantity"] * item["unit_price"])

    tmp_dir = tempfile.mkdtemp()
    try:
        tmp_xlsx = os.path.join(tmp_dir, "invoice.xlsx")
        wb.save(tmp_xlsx)
        pdf_dir = os.path.join(tmp_dir, "pdf")
        os.makedirs(pdf_dir, exist_ok=True)
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", pdf_dir, tmp_xlsx],
            check=True,
            timeout=120,
        )
        tmp_pdf = os.path.join(pdf_dir, "invoice.pdf")
        shamsi_clean = data["date_shamsi"].replace("/", "")
        year_month = shamsi_clean[:6]
        month_dir = OUTPUT_DIR / year_month
        month_dir.mkdir(parents=True, exist_ok=True)
        safe_name = data["customer_name"].replace(" ", "_")
        final_name = f"{data['invoice_number']}-{shamsi_clean}-{safe_name}.pdf"
        final_path = month_dir / final_name
        shutil.move(tmp_pdf, final_path)
        return str(final_path)
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

def get_existing_pdf(invoice_number, date_shamsi):
    shamsi_clean = date_shamsi.replace("/", "")
    year_month = shamsi_clean[:6]
    month_dir = OUTPUT_DIR / year_month
    if month_dir.exists():
        for f in month_dir.glob(f"{invoice_number}-{shamsi_clean}-*.pdf"):
            return str(f)
    return None

# ---------- Keyboards ----------
def main_menu_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 فاکتور جدید", callback_data="new_invoice")],
        [InlineKeyboardButton("📂 فاکتورهای سابق", callback_data="old_invoices")],
        [InlineKeyboardButton("👥 لیست مشتریان", callback_data="customer_list")],
    ])

def back_to_main_kb():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🔙 بازگشت به منوی اصلی", callback_data="back_to_main")]
    ])

# ---------- Handlers ----------
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "خوش آمدید. لطفاً یک گزینه را انتخاب کنید:",
        reply_markup=main_menu_kb(),
    )
    return State.MAIN_MENU

async def main_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "new_invoice":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("مشتری جدید", callback_data="new_customer")],
            [InlineKeyboardButton("مشتری قبلی", callback_data="existing_customer")],
            [InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_main")],
        ])
        await query.edit_message_text("آیا مشتری جدید است یا مشتری قبلی؟", reply_markup=kb)
        return State.NEW_INVOICE_MENU

    if data == "old_invoices":
        await query.edit_message_text(
            "نام مشتری را برای جستجو وارد کنید (یا * برای همه):",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_main")]
            ]),
        )
        return State.ENTER_SEARCH_NAME

    if data == "customer_list":
        await query.edit_message_text(
            "نام مشتری را برای جستجو وارد کنید (یا * برای همه):",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_main")]
            ]),
        )
        context.user_data["list_mode"] = "customer"
        return State.CUSTOMER_LIST

    if data == "back_to_main":
        context.user_data.clear()
        await query.edit_message_text("منوی اصلی:", reply_markup=main_menu_kb())
        return State.MAIN_MENU

# ---- New invoice flow ----
async def new_invoice_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data

    if data == "new_customer":
        context.user_data["invoice"] = {"is_new": True, "items": []}
        await query.edit_message_text("نام و نام خانوادگی مشتری جدید را وارد کنید:")
        return State.ENTER_NEW_CUSTOMER_NAME

    if data == "existing_customer":
        context.user_data["invoice"] = {"is_new": False, "items": []}
        await query.edit_message_text("نام مشتری قبلی را برای جستجو وارد کنید:")
        return State.ENTER_SEARCH_NAME

    if data == "back_to_main":
        context.user_data.clear()
        await query.edit_message_text("منوی اصلی:", reply_markup=main_menu_kb())
        return State.MAIN_MENU

async def enter_new_customer_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("نام نمی‌تواند خالی باشد. دوباره وارد کنید:")
        return State.ENTER_NEW_CUSTOMER_NAME
    context.user_data["invoice"]["customer_name"] = name
    await update.message.reply_text("استان را وارد کنید:")
    return State.ENTER_NEW_CUSTOMER_STATE

async def enter_new_customer_state(update: Update, context: ContextTypes.DEFAULT_TYPE):
    state = update.message.text.strip()
    if not state:
        await update.message.reply_text("استان نمی‌تواند خالی باشد. دوباره وارد کنید:")
        return State.ENTER_NEW_CUSTOMER_STATE
    context.user_data["invoice"]["customer_state"] = state
    await update.message.reply_text("شهر را وارد کنید:")
    return State.ENTER_NEW_CUSTOMER_CITY

async def enter_new_customer_city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    city = update.message.text.strip()
    if not city:
        await update.message.reply_text("شهر نمی‌تواند خالی باشد. دوباره وارد کنید:")
        return State.ENTER_NEW_CUSTOMER_CITY
    context.user_data["invoice"]["customer_city"] = city
    await update.message.reply_text("نشانی را وارد کنید:")
    return State.ENTER_NEW_CUSTOMER_ADDRESS

async def enter_new_customer_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    address = update.message.text.strip()
    if not address:
        await update.message.reply_text("نشانی نمی‌تواند خالی باشد. دوباره وارد کنید:")
        return State.ENTER_NEW_CUSTOMER_ADDRESS
    context.user_data["invoice"]["customer_address"] = address
    await update.message.reply_text("شماره تلفن مشتری را وارد کنید:")
    return State.ENTER_NEW_CUSTOMER_PHONE

async def enter_new_customer_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    phone = update.message.text.strip()
    if not phone:
        await update.message.reply_text("شماره تلفن نمی‌تواند خالی باشد. دوباره وارد کنید:")
        return State.ENTER_NEW_CUSTOMER_PHONE
    context.user_data["invoice"]["customer_phone"] = phone
    inv = context.user_data["invoice"]
    cust_id = create_customer(
        inv["customer_name"],
        inv["customer_state"],
        inv["customer_city"],
        inv["customer_address"],
        inv["customer_phone"],
    )
    inv["customer_id"] = cust_id
    inv["invoice_number"] = get_next_invoice_number()
    await update.message.reply_text(
        "تاریخ فاکتور را به صورت شمسی وارد کنید (مثال: 1405/06/06):"
    )
    return State.ENTER_DATE_SHAMSI

async def enter_date_shamsi(update: Update, context: ContextTypes.DEFAULT_TYPE):
    date_text = update.message.text.strip()
    parts = date_text.split("/")
    if len(parts) != 3 or not all(p.isdigit() for p in parts):
        await update.message.reply_text("تاریخ نامعتبر است. مثال: 1405/06/06")
        return State.ENTER_DATE_SHAMSI
    context.user_data["invoice"]["date_shamsi"] = date_text
    context.user_data["current_item"] = {}
    await update.message.reply_text("کالای اول را وارد کنید - شرح کالا چیست؟")
    return State.ENTER_ITEM_DESC

async def enter_search_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    name = update.message.text.strip()
    if not name:
        await update.message.reply_text("نام نمی‌تواند خالی باشد.")
        return State.ENTER_SEARCH_NAME
    rows = search_customers(name)
    if not rows:
        await update.message.reply_text(
            "مشتری یافت نشد. نام دیگری وارد کنید یا * برای همه بزنید.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_main")]
            ]),
        )
        return State.ENTER_SEARCH_NAME
    context.user_data["search_results"] = rows
    context.user_data["search_page"] = 0
    await show_customer_page(update, context)
    return State.SELECT_EXISTING_CUSTOMER

async def show_customer_page(update_or_query, context):
    results = context.user_data.get("search_results", [])
    page = context.user_data.get("search_page", 0)
    start = page * ITEMS_PER_PAGE
    end = start + ITEMS_PER_PAGE
    page_rows = results[start:end]
    total_pages = (len(results) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    kb = [
        [InlineKeyboardButton(f"{r[1]} - {r[5]}", callback_data=f"pickcust_{r[0]}")]
        for r in page_rows
    ]
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ قبلی", callback_data="search_page_-1"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("بعدی ▶️", callback_data="search_page_+1"))
    if nav:
        kb.append(nav)
    kb.append([InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_main")])
    text = f"نتایج جستجو ({len(results)} مشتری) - صفحه {page+1}/{max(total_pages,1)}:\nیک مشتری را انتخاب کنید:"
    if isinstance(update_or_query, Update):
        await update_or_query.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await update_or_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))

async def select_existing_customer(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "back_to_main":
        context.user_data.clear()
        await query.edit_message_text("منوی اصلی:", reply_markup=main_menu_kb())
        return State.MAIN_MENU
    if data.startswith("search_page_"):
        direction = data.split("_")[-1]
        page = context.user_data.get("search_page", 0)
        if direction == "+1":
            page += 1
        else:
            page = max(page - 1, 0)
        context.user_data["search_page"] = page
        await show_customer_page(query, context)
        return State.SELECT_EXISTING_CUSTOMER
    if data.startswith("pickcust_"):
        cust_id = int(data.split("_", 1)[1])
        cust = get_customer(cust_id)
        if not cust:
            await query.edit_message_text("مشتری یافت نشد.", reply_markup=back_to_main_kb())
            return State.MAIN_MENU
        inv = context.user_data["invoice"]
        inv["customer_id"] = cust["id"]
        inv["customer_name"] = cust["name"]
        inv["customer_state"] = cust["state"]
        inv["customer_city"] = cust["city"]
        inv["customer_address"] = cust["address"]
        inv["customer_phone"] = cust["phone"]
        inv["invoice_number"] = get_next_invoice_number()
        await query.edit_message_text(
            f"مشتری انتخاب شد: {cust['name']}\n\n"
            f"تاریخ فاکتور را به صورت شمسی وارد کنید (مثال: 1405/06/06):"
        )
        return State.ENTER_DATE_SHAMSI
    return State.SELECT_EXISTING_CUSTOMER

# ---- Items flow ----
async def enter_item_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    desc = update.message.text.strip()
    if not desc:
        await update.message.reply_text("شرح کالا نمی‌تواند خالی باشد.")
        return State.ENTER_ITEM_DESC
    context.user_data["current_item"] = {"description": desc}
    await update.message.reply_text("تعداد کالا را وارد کنید (مثال: 5):")
    return State.ENTER_ITEM_QTY

async def enter_item_qty(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    try:
        qty = float(text)
        if qty <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("عدد معتبر وارد کنید (مثلاً 5 یا 2.5):")
        return State.ENTER_ITEM_QTY
    context.user_data["current_item"]["quantity"] = qty
    await update.message.reply_text("واحد کالا را وارد کنید (مثال: عدد، کیلو، متر):")
    return State.ENTER_ITEM_UNIT

async def enter_item_unit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    unit = update.message.text.strip()
    if not unit:
        await update.message.reply_text("واحد نمی‌تواند خالی باشد.")
        return State.ENTER_ITEM_UNIT
    context.user_data["current_item"]["unit"] = unit
    await update.message.reply_text("قیمت واحد را به ریال وارد کنید (مثال: 150000):")
    return State.ENTER_ITEM_PRICE

async def enter_item_price(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().replace(",", "").replace("،", "")
    try:
        price = int(text)
        if price < 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("عدد معتبر وارد کنید (مثال: 150000):")
        return State.ENTER_ITEM_PRICE
    item = context.user_data["current_item"]
    item["unit_price"] = price
    item["total_price"] = int(item["quantity"] * price)
    context.user_data["invoice"]["items"].append(item)
    context.user_data["current_item"] = {}
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ افزودن کالای بعدی", callback_data="add_more")],
        [InlineKeyboardButton("✅ پایان و بررسی", callback_data="finish_items")],
        [InlineKeyboardButton("❌ انصراف", callback_data="cancel_invoice")],
    ])
    await update.message.reply_text(
        f"کالا ثبت شد: {item['description']} - {item['quantity']} {item['unit']} - {item['total_price']:,} ریال\n\n"
        "می‌خواهید کالای دیگری اضافه کنید؟",
        reply_markup=kb,
    )
    return State.ASK_MORE_ITEMS

async def ask_more_items(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "add_more":
        await query.edit_message_text("شرح کالای جدید را وارد کنید:")
        return State.ENTER_ITEM_DESC
    if query.data == "cancel_invoice":
        context.user_data.clear()
        await query.edit_message_text("فاکتور لغو شد.", reply_markup=main_menu_kb())
        return State.MAIN_MENU
    inv = context.user_data["invoice"]
    items = inv["items"]
    lines = []
    total = 0
    for i, it in enumerate(items, 1):
        line_total = int(it["quantity"] * it["unit_price"])
        total += line_total
        lines.append(
            f"{i}. {it['description']} | {it['quantity']} {it['unit']} | "
            f"{it['unit_price']:,} ریال | {line_total:,} ریال"
        )
    inv["total"] = total
    msg = (
        f"📋 پیش‌نمایش فاکتور:\n\n"
        f"شماره فاکتور: {inv['invoice_number']}\n"
        f"تاریخ: {inv['date_shamsi']}\n"
        f"خریدار: {inv['customer_name']} - {inv['customer_phone']}\n"
        f"استان: {inv.get('customer_state','')}  شهر: {inv.get('customer_city','')}\n"
        f"نشانی: {inv.get('customer_address','')}\n\n"
        f"کالاها:\n" + "\n".join(lines) + f"\n\n💰 جمع کل: {total:,} ریال\n\n"
        "آیا تأیید می‌کنید؟"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ تأیید و ارسال PDF", callback_data="confirm_invoice")],
        [InlineKeyboardButton("✏️ ویرایش کالاها", callback_data="edit_items")],
        [InlineKeyboardButton("❌ انصراف", callback_data="cancel_invoice")],
    ])
    await query.edit_message_text(msg, reply_markup=kb)
    return State.REVIEW_INVOICE

async def review_invoice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "cancel_invoice":
        context.user_data.clear()
        await query.edit_message_text("فاکتور لغو شد.", reply_markup=main_menu_kb())
        return State.MAIN_MENU
    if data == "edit_items":
        context.user_data["invoice"]["items"] = []
        context.user_data["current_item"] = {}
        await query.edit_message_text("کالاها پاک شدند. شرح کالای اول را وارد کنید:")
        return State.ENTER_ITEM_DESC
    inv = context.user_data["invoice"]
    pdf_data = {
        "customer_name": inv["customer_name"],
        "customer_phone": inv["customer_phone"],
        "customer_state": inv.get("customer_state", ""),
        "customer_city": inv.get("customer_city", ""),
        "customer_address": inv.get("customer_address", ""),
        "invoice_number": inv["invoice_number"],
        "date_shamsi": inv["date_shamsi"],
        "items": inv["items"],
    }
    try:
        await query.edit_message_text("در حال ساخت PDF...")
        pdf_path = generate_invoice_pdf(pdf_data)
    except subprocess.TimeoutExpired:
        await query.edit_message_text("تبدیل PDF بیش از حد طول کشید. دوباره تلاش کنید.")
        return State.REVIEW_INVOICE
    except Exception as e:
        logger.exception("PDF generation failed")
        await query.edit_message_text(f"خطا در ساخت PDF: {e}")
        return State.REVIEW_INVOICE

    total = sum(int(it["quantity"] * it["unit_price"]) for it in inv["items"])
    inv_id = save_invoice(
        inv["invoice_number"], inv["date_shamsi"],
        inv["customer_id"], total, pdf_path,
    )
    save_items(inv_id, inv["items"])
    with open(pdf_path, "rb") as f:
        await context.bot.send_document(
            chat_id=query.message.chat_id,
            document=f,
            filename=os.path.basename(pdf_path),
        )
    context.user_data.clear()
    await query.message.reply_text(
        "✅ فاکتور با موفقیت ساخته و ارسال شد.",
        reply_markup=main_menu_kb(),
    )
    return State.MAIN_MENU

# ---- Customer list & pagination ----
async def show_customer_list_page(update_or_query, context):
    mode = context.user_data.get("list_mode", "customer")
    if mode == "customer":
        query_text = context.user_data.get("list_search", "")
        if query_text == "" or query_text == "*":
            conn = get_conn()
            try:
                with conn.cursor() as cur:
                    cur.execute("SELECT id, name, state, city, address, phone FROM customers ORDER BY name")
                    rows = cur.fetchall()
            finally:
                conn.close()
        else:
            rows = search_customers(query_text)
        title = "لیست مشتریان"
        item_txt = lambda r: f"{r[1]} - {r[5]}"
        cb_prefix = "viewcust"
    else:
        cust_id = context.user_data.get("hist_cust_id")
        rows = get_customer_invoices(cust_id)
        title = f"تاریخچه فاکتورهای مشتری: {context.user_data.get('hist_cust_name','')}"
        item_txt = lambda r: f"#{r[0]} - {r[1]} - {r[2]:,} ریال"
        cb_prefix = "viewinv"

    page = context.user_data.get("list_page", 0)
    start = page * ITEMS_PER_PAGE
    end = start + ITEMS_PER_PAGE
    page_rows = rows[start:end]
    total_pages = (len(rows) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE
    kb = [
        [InlineKeyboardButton(item_txt(r), callback_data=f"{cb_prefix}_{r[0]}")]
        for r in page_rows
    ]
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ قبلی", callback_data="list_page_-1"))
    if page < total_pages - 1:
        nav.append(InlineKeyboardButton("بعدی ▶️", callback_data="list_page_+1"))
    if nav:
        kb.append(nav)
    kb.append([InlineKeyboardButton("🔙 بازگشت", callback_data="back_to_main")])
    text = f"{title} ({len(rows)} مورد) - صفحه {page+1}/{max(total_pages,1)}:\nیک مورد را انتخاب کنید:"
    if isinstance(update_or_query, Update):
        await update_or_query.message.reply_text(text, reply_markup=InlineKeyboardMarkup(kb))
    else:
        await update_or_query.edit_message_text(text, reply_markup=InlineKeyboardMarkup(kb))

async def customer_list_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.data == "back_to_main":
        context.user_data.clear()
        await query.edit_message_text("منوی اصلی:", reply_markup=main_menu_kb())
        return State.MAIN_MENU
    if query.data.startswith("list_page_"):
        direction = query.data.split("_")[-1]
        page = context.user_data.get("list_page", 0)
        if direction == "+1":
            page += 1
        else:
            page = max(page - 1, 0)
        context.user_data["list_page"] = page
        await show_customer_list_page(query, context)
        return State.CUSTOMER_PAGE
    if query.data.startswith("viewcust_"):
        cust_id = int(query.data.split("_", 1)[1])
        cust = get_customer(cust_id)
        if not cust:
            await query.edit_message_text("مشتری یافت نشد.", reply_markup=back_to_main_kb())
            return State.MAIN_MENU
        context.user_data["viewing_customer_id"] = cust_id
        await view_customer_handler(update, context)
        return State.VIEW_CUSTOMER
    if query.data.startswith("viewinv_"):
        inv_id = int(query.data.split("_", 1)[1])
        conn = get_conn()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT invoice_number, date_shamsi, file_path FROM invoices WHERE id=%s",
                    (inv_id,),
                )
                row = cur.fetchone()
        finally:
            conn.close()
        if not row:
            await query.edit_message_text("فاکتور یافت نشد.", reply_markup=back_to_main_kb())
            return State.MAIN_MENU
        inv_number, date_shamsi, file_path = row
        # Prefer existing PDF
        if not file_path or not os.path.exists(file_path):
            existing = get_existing_pdf(inv_number, date_shamsi)
            if existing:
                file_path = existing
                conn = get_conn()
                try:
                    with conn.cursor() as cur:
                        cur.execute(
                            "UPDATE invoices SET file_path=%s WHERE id=%s",
                            (file_path, inv_id),
                        )
                        conn.commit()
                finally:
                    conn.close()
            else:
                # Regenerate from DB
                conn = get_conn()
                try:
                    with conn.cursor() as cur:
                        cur.execute("SELECT customer_id FROM invoices WHERE id=%s", (inv_id,))
                        cust_row = cur.fetchone()
                        cust_id = cust_row[0] if cust_row else None
                        cust = get_customer(cust_id) if cust_id else None
                        if cust:
                            cur.execute(
                                "SELECT description, quantity, unit, unit_price FROM invoice_items WHERE invoice_id=%s",
                                (inv_id,),
                            )
                            items_rows = cur.fetchall()
                            items = [
                                {"description": d, "quantity": q, "unit": u, "unit_price": up}
                                for d, q, u, up in items_rows
                            ]
                            pdf_data = {
                                "customer_name": cust["name"],
                                "customer_phone": cust["phone"],
                                "customer_state": cust.get("state",""),
                                "customer_city": cust.get("city",""),
                                "customer_address": cust.get("address",""),
                                "invoice_number": inv_number,
                                "date_shamsi": date_shamsi,
                                "items": items,
                            }
                            file_path = generate_invoice_pdf(pdf_data)
                            with conn.cursor() as cur:
                                cur.execute(
                                    "UPDATE invoices SET file_path=%s WHERE id=%s",
                                    (file_path, inv_id),
                                )
                                conn.commit()
                finally:
                    conn.close()
        try:
            with open(file_path, "rb") as f:
                await context.bot.send_document(
                    chat_id=query.message.chat_id,
                    document=f,
                    filename=os.path.basename(file_path),
                )
            await query.message.reply_text("PDF ارسال شد.")
        except Exception as e:
            await query.edit_message_text(f"خطا در ارسال PDF: {e}", reply_markup=back_to_main_kb())
        return State.MAIN_MENU
    return State.CUSTOMER_PAGE

async def view_customer_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "back_to_main":
        context.user_data.clear()
        await query.edit_message_text("منوی اصلی:", reply_markup=main_menu_kb())
        return State.MAIN_MENU
    if data.startswith("viewcust_"):
        cust_id = int(data.split("_", 1)[1])
        cust = get_customer(cust_id)
        if not cust:
            await query.edit_message_text("مشتری یافت نشد.", reply_markup=back_to_main_kb())
            return State.MAIN_MENU
        context.user_data["viewing_customer_id"] = cust_id
        msg = (
            f"👤 اطلاعات مشتری:\n\n"
            f"نام: {cust['name']}\n"
            f"استان: {cust['state']}\n"
            f"شهر: {cust['city']}\n"
            f"نشانی: {cust['address']}\n"
            f"تلفن: {cust['phone']}\n"
        )
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("✏️ ویرایش نام", callback_data="edit_name")],
            [InlineKeyboardButton("✏️ ویرایش استان", callback_data="edit_state")],
            [InlineKeyboardButton("✏️ ویرایش شهر", callback_data="edit_city")],
            [InlineKeyboardButton("✏️ ویرایش نشانی", callback_data="edit_address")],
            [InlineKeyboardButton("✏️ ویرایش تلفن", callback_data="edit_phone")],
            [InlineKeyboardButton("📂 تاریخچه فاکتورها", callback_data="history_invoices")],
            [InlineKeyboardButton("🔙 بازگشت به لیست", callback_data="back_to_list")],
        ])
        await query.edit_message_text(msg, reply_markup=kb)
        return State.EDIT_CUSTOMER_FIELD
    return State.VIEW_CUSTOMER

async def customer_edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    if data == "back_to_list":
        context.user_data["list_mode"] = "customer"
        context.user_data["list_search"] = ""
        context.user_data["list_page"] = 0
        await show_customer_list_page(query, context)
        return State.CUSTOMER_LIST
    if data == "history_invoices":
        cust_id = context.user_data.get("viewing_customer_id")
        cust = get_customer(cust_id)
        if not cust:
            await query.edit_message_text("مشتری یافت نشد.", reply_markup=back_to_main_kb())
            return State.MAIN_MENU
        context.user_data["hist_cust_id"] = cust_id
        context.user_data["hist_cust_name"] = cust["name"]
        context.user_data["list_mode"] = "history"
        context.user_data["list_page"] = 0
        await show_customer_list_page(query, context)
        return State.CUSTOMER_PAGE
    field_map = {
        "edit_name": "name",
        "edit_state": "state",
        "edit_city": "city",
        "edit_address": "address",
        "edit_phone": "phone",
    }
    field = field_map.get(data)
    if not field:
        return State.EDIT_CUSTOMER_FIELD
    context.user_data["edit_field"] = field
    cust_id = context.user_data.get("viewing_customer_id")
    cust = get_customer(cust_id)
    cur = cust.get(field, "")
    await query.edit_message_text(f"مقدار جدید برای {field} را وارد کنید (فعلی: {cur}):")
    return State.EDIT_CUSTOMER_VALUE

async def customer_save_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    new_value = update.message.text.strip()
    cust_id = context.user_data.get("viewing_customer_id")
    field = context.user_data.get("edit_field")
    if not cust_id or not field:
        await update.message.reply_text("خطا. دوباره شروع کنید.", reply_markup=main_menu_kb())
        return State.MAIN_MENU
    if not new_value:
        await update.message.reply_text("مقدار نمی‌تواند خالی باشد. دوباره وارد کنید:")
        return State.EDIT_CUSTOMER_VALUE
    cust = get_customer(cust_id)
    if field == "name":
        update_customer(cust_id, new_value, cust["state"], cust["city"], cust["address"], cust["phone"])
    elif field == "state":
        update_customer(cust_id, cust["name"], new_value, cust["city"], cust["address"], cust["phone"])
    elif field == "city":
        update_customer(cust_id, cust["name"], cust["state"], new_value, cust["address"], cust["phone"])
    elif field == "address":
        update_customer(cust_id, cust["name"], cust["state"], cust["city"], new_value, cust["phone"])
    elif field == "phone":
        update_customer(cust_id, cust["name"], cust["state"], cust["city"], cust["address"], new_value)
    cust = get_customer(cust_id)
    msg = (
        f"👤 اطلاعات مشتری (به‌روز شد):\n\n"
        f"نام: {cust['name']}\n"
        f"استان: {cust['state']}\n"
        f"شهر: {cust['city']}\n"
        f"نشانی: {cust['address']}\n"
        f"تلفن: {cust['phone']}\n"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ ویرایش نام", callback_data="edit_name")],
        [InlineKeyboardButton("✏️ ویرایش استان", callback_data="edit_state")],
        [InlineKeyboardButton("✏️ ویرایش شهر", callback_data="edit_city")],
        [InlineKeyboardButton("✏️ ویرایش نشانی", callback_data="edit_address")],
        [InlineKeyboardButton("✏️ ویرایش تلفن", callback_data="edit_phone")],
        [InlineKeyboardButton("📂 تاریخچه فاکتورها", callback_data="history_invoices")],
        [InlineKeyboardButton("🔙 بازگشت به لیست", callback_data="back_to_list")],
    ])
    await update.message.reply_text(msg, reply_markup=kb)
    return State.VIEW_CUSTOMER

# ---- Cancel ----
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("عملیات لغو شد.", reply_markup=main_menu_kb())
    return State.MAIN_MENU

# ---------- Main ----------
def main():
    init_db()
    token = os.environ.get("TELEGRAM_BOT_TOKEN")
    if not token:
        raise RuntimeError("TELEGRAM_BOT_TOKEN env var not set")
    app = Application.builder().token(token).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
            State.MAIN_MENU: [CallbackQueryHandler(main_menu_handler)],
            State.NEW_INVOICE_MENU: [CallbackQueryHandler(new_invoice_menu)],
            State.ENTER_NEW_CUSTOMER_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_new_customer_name)
            ],
            State.ENTER_NEW_CUSTOMER_STATE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_new_customer_state)
            ],
            State.ENTER_NEW_CUSTOMER_CITY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_new_customer_city)
            ],
            State.ENTER_NEW_CUSTOMER_ADDRESS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_new_customer_address)
            ],
            State.ENTER_NEW_CUSTOMER_PHONE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_new_customer_phone)
            ],
            State.ENTER_SEARCH_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_search_name)
            ],
            State.SELECT_EXISTING_CUSTOMER: [
                CallbackQueryHandler(select_existing_customer)
            ],
            State.ENTER_DATE_SHAMSI: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_date_shamsi)
            ],
            State.ENTER_ITEM_DESC: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_item_desc)
            ],
            State.ENTER_ITEM_QTY: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_item_qty)
            ],
            State.ENTER_ITEM_UNIT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_item_unit)
            ],
            State.ENTER_ITEM_PRICE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, enter_item_price)
            ],
            State.ASK_MORE_ITEMS: [CallbackQueryHandler(ask_more_items)],
            State.REVIEW_INVOICE: [CallbackQueryHandler(review_invoice_handler)],
            State.CUSTOMER_LIST: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, customer_list_handler)
            ],
            State.CUSTOMER_PAGE: [
                CallbackQueryHandler(customer_list_handler)
            ],
            State.VIEW_CUSTOMER: [
                CallbackQueryHandler(view_customer_handler)
            ],
            State.EDIT_CUSTOMER_FIELD: [
                CallbackQueryHandler(customer_edit_field)
            ],
            State.EDIT_CUSTOMER_VALUE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, customer_save_edit)
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
    )
    app.add_handler(conv)
    app.run_polling()

if __name__ == "__main__":
    main()
